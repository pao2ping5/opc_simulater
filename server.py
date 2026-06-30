"""
OPC UA Simulator Server (CLI mode)
═══════════════════════════════════
Reads a model Excel, creates OPC UA address space, and periodically
updates values using the pluggable strategy engine.  Shares the same
:class:`GenericOPCSimulator` as ``web_server.py`` — no more diverging
copies.

Usage:  python server.py [--xlsx path/to/model.xlsx]
"""

from __future__ import annotations

import logging
import os
import signal
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from common import (
    NodeMeta,
    build_device_model,
    read_model_excel,
)
from simulator import DEFAULT_OPC_ENDPOINT, GenericOPCSimulator

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = SCRIPT_DIR.parent / "opc_list_test.xlsx"
UPDATE_INTERVAL = 2  # seconds

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger("opc-sim-server")


# ========================================================================
# Point table generator
# ========================================================================


def generate_reading_excel(
    model_nodes: list[NodeMeta],
    output_path: str | Path,
    ns: int = 2,
) -> None:
    """Generate an OPC client-side point table for the simulator.

    Writes an Excel with columns: [name, NodeId address] for each node,
    suitable for importing into an OPC client.
    """
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "PointTable"

    # Header
    ws.cell(row=1, column=1, value="名称")
    ws.cell(row=1, column=2, value="NodeId")

    for i, meta in enumerate(model_nodes, start=2):
        ws.cell(row=i, column=1, value=meta.effective_display_name)
        ws.cell(row=i, column=2, value=f"ns={ns};s={meta.node_id}")

    wb.save(str(output_path))
    log.info("Client point table written to %s (%d rows)", output_path, len(model_nodes))


# ========================================================================
# Entry point
# ========================================================================

if __name__ == "__main__":
    xlsx_path = (
        sys.argv[1] if len(sys.argv) > 1 else os.environ.get("OPC_POINT_TABLE", str(DEFAULT_XLSX))
    )
    output_path = SCRIPT_DIR / "opc_sim_list.xlsx"

    log.info("Loading model from %s", xlsx_path)
    nodes, warnings = read_model_excel(xlsx_path)
    for w in warnings:
        log.warning("  %s", w)

    model = build_device_model(nodes)
    log.info("Parsed %d nodes across %d groups", len(model.nodes), len(model.groups))

    # Generate client point table
    all_nodes = list(model.nodes.values())
    generate_reading_excel(all_nodes, output_path)

    # Start simulator — CLI mode: no state_file (no persistence needed)
    sim = GenericOPCSimulator(DEFAULT_OPC_ENDPOINT)
    sim.setup(model)

    def _shutdown(signum: int, frame: Any) -> None:
        log.info("Received signal %d, shutting down...", signum)
        sim.running = False

    signal.signal(signal.SIGINT, _shutdown)
    signal.signal(signal.SIGTERM, _shutdown)

    sim.run(tick_interval=UPDATE_INTERVAL)
