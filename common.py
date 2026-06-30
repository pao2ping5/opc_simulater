"""
OPC Simulator — Generic model definition, strategy registry, and value engine.

No industry-specific knowledge. Everything is driven by the model definition
imported from an Excel spreadsheet and further tuned through the Web UI.
"""

from __future__ import annotations

import json
import math
import os
import random
import tempfile
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

from openpyxl import load_workbook

# ============================================================================
# Data models
# ============================================================================


@dataclass
class NodeMeta:
    """Complete metadata for a single OPC UA variable node.

    The ``node_id`` encodes the address-space tree via dot-separated segments.
    For example ``Shearer.LeftMotor.Current`` produces::

        Objects/
          Shearer (Folder)
            LeftMotor (Folder)
              Current (Variable)

    Parameters
    ----------
    node_id:
        Dot-separated path, e.g. ``"Shearer.LeftMotor.Current"``.
    data_type:
        OPC UA data type.  One of ``"float"``, ``"int"``, ``"bool"``, ``"string"``.
    range_lo / range_hi:
        Value range used for EURange and simulation clamping.
    unit:
        Engineering unit string (e.g. ``"A"``, ``"MPa"``, ``"rpm"``).
    gen_strategy:
        Name of the value-generation strategy (see :class:`StrategyRegistry`).
    gen_params:
        Strategy-specific parameters as a flat dict (e.g. ``{"center_ratio": 0.6}``).
    description:
        Free-text description, mapped to OPC UA ``Description`` attribute.
    instance_count:
        How many concrete OPC nodes this row represents.  Instance ``i`` gets
        a NodeId of ``node_id + "_" + str(i).zfill(3)``.
    display_name:
        Optional OPC UA ``DisplayName`` override.  When empty the leaf segment
        of ``node_id`` is used.
    group_depth:
        How many dot-separated segments of ``node_id`` define the equipment
        group.  Default ``1`` → ``"Shearer"`` is the group; ``2`` →
        ``"Shearer.LeftMotor"`` would be the group.
    """

    node_id: str
    data_type: str = "float"
    range_lo: float = 0.0
    range_hi: float = 100.0
    unit: str = ""
    gen_strategy: str = ""
    gen_params: Dict[str, Any] = field(default_factory=dict)
    description: str = ""
    instance_count: int = 1
    display_name: str = ""
    group_depth: int = 1

    # ---- derived helpers --------------------------------------------------

    @property
    def parts(self) -> List[str]:
        return self.node_id.split(".")

    @property
    def group_key(self) -> str:
        d = self.group_depth if self.group_depth > 0 else 1
        return ".".join(self.parts[: min(d, len(self.parts))])

    @property
    def leaf_name(self) -> str:
        return self.parts[-1] if self.parts else self.node_id

    @property
    def parent_path(self) -> str:
        return ".".join(self.parts[:-1])

    @property
    def effective_display_name(self) -> str:
        return self.display_name or self.leaf_name

    def instance_node_id(self, idx: int) -> str:
        """``node_id`` with zero-padded instance suffix."""
        return f"{self.node_id}_{idx:03d}"


@dataclass
class GroupInfo:
    """Summary of a node-id prefix group, used by the Web UI for tabs."""

    key: str
    label: str
    node_ids: List[str] = field(default_factory=list)
    node_count: int = 0


@dataclass
class DeviceModel:
    """A complete parsed device model ready to be served as an OPC UA simulation."""

    nodes: Dict[str, NodeMeta] = field(default_factory=dict)
    """``node_id`` → ``NodeMeta`` (after instance expansion)."""

    groups: Dict[str, GroupInfo] = field(default_factory=dict)
    """Group key → GroupInfo (derived from node_id prefixes)."""

    separator: str = "."
    """Separator used in node_ids."""


# ============================================================================
# Strategy registry – pluggable value generation
# ============================================================================


@dataclass
class ValueStrategy:
    """A named value-generation strategy with its own logic.

    The callable receives ``(lo, hi, params, current_val, elapsed, tick)``
    and returns the next raw numeric value.
    """

    name: str
    description: str
    generate: Callable[..., float]


# -- built-in strategy implementations ---------------------------------------


def _make_uniform(
    lo: float, hi: float, _p: dict, _cur: float, _e: float, _t: int
) -> float:
    return lo + (hi - lo) * random.random()


def _make_normal(
    lo: float, hi: float, p: dict, _cur: float, _e: float, _t: int
) -> float:
    center = lo + (hi - lo) * p.get("center_ratio", 0.5)
    std = (hi - lo) * p.get("std_ratio", 0.1)
    return max(lo, min(hi, random.gauss(center, std)))


def _make_sinusoidal(
    lo: float, hi: float, p: dict, _cur: float, elapsed: float, _t: int
) -> float:
    center = (lo + hi) / 2
    amplitude = (hi - lo) / 2
    period = p.get("period_sec", 60.0)
    phase = p.get("phase", 0.0)
    return center + amplitude * math.sin(2 * math.pi * elapsed / period + phase)


def _make_binary(
    _lo: float, _hi: float, p: dict, _cur: float, _e: float, _t: int
) -> float:
    return float(random.randint(0, 1))


def _make_random_walk(
    lo: float, hi: float, p: dict, cur: float, _e: float, _t: int
) -> float:
    step = (hi - lo) * p.get("step_ratio", 0.02)
    val = cur + (2 * random.random() - 1) * step
    return max(lo, min(hi, val))


def _make_ramp(
    lo: float, hi: float, p: dict, _cur: float, _e: float, tick: int
) -> float:
    step = (hi - lo) * p.get("step_ratio", 0.01)
    # Guard against zero/negative step (hi==lo or step_ratio=0) → division by zero
    if step <= 0:
        return lo
    direction = p.get("direction", 1.0)
    span = int((hi - lo) / step + 1)
    if span <= 0:
        return lo
    cur_calc = lo + (tick % span) * step * direction
    if direction > 0 and cur_calc > hi:
        cur_calc = lo
    elif direction < 0 and cur_calc < lo:
        cur_calc = hi
    return cur_calc


def _make_counter(
    lo: float, hi: float, p: dict, _cur: float, _e: float, tick: int
) -> float:
    step = p.get("step", 1.0)
    # Guard: modulo by zero when hi - lo + step == 0
    modulus = hi - lo + step
    if modulus <= 0:
        return lo
    return lo + ((tick * step) % modulus)


def _make_constant(
    _lo: float, _hi: float, p: dict, _cur: float, _e: float, _t: int
) -> float:
    return float(p.get("value", 0.0))


def _make_current(
    lo: float, hi: float, p: dict, _cur: float, _e: float, _t: int
) -> float:
    center = lo + (hi - lo) * p.get("center_ratio", 0.6)
    jitter = (hi - lo) * p.get("jitter_ratio", 0.15)
    return max(lo, min(hi, center + jitter * (2 * random.random() - 1)))


def _make_temp(lo: float, hi: float, p: dict, _cur: float, _e: float, _t: int) -> float:
    center = lo + (hi - lo) * p.get("center_ratio", 0.4)
    jitter = (hi - lo) * p.get("jitter_ratio", 0.1)
    return max(lo, min(hi, center + jitter * (2 * random.random() - 1)))


def _make_pressure(
    lo: float, hi: float, p: dict, _cur: float, _e: float, _t: int
) -> float:
    center = lo + (hi - lo) * p.get("center_ratio", 0.5)
    jitter = (hi - lo) * p.get("jitter_ratio", 0.05)
    return max(lo, min(hi, center + jitter * (2 * random.random() - 1)))


def _make_speed(
    lo: float, hi: float, p: dict, _cur: float, _e: float, _t: int
) -> float:
    center = lo + (hi - lo) * p.get("center_ratio", 0.7)
    jitter = (hi - lo) * p.get("jitter_ratio", 0.05)
    return max(lo, min(hi, center + jitter * (2 * random.random() - 1)))


_BUILTIN_STRATEGIES: List[ValueStrategy] = [
    ValueStrategy(
        "random_uniform", "均匀随机，在 [lo, hi] 范围内均匀分布", _make_uniform
    ),
    ValueStrategy(
        "random_normal",
        "正态分布，中心=(lo+hi)*center_ratio，std=(hi-lo)*std_ratio",
        _make_normal,
    ),
    ValueStrategy("binary_toggle", "随机 0/1 翻转，用于布尔信号", _make_binary),
    ValueStrategy(
        "sinusoidal", "正弦波扫全量程，period_sec 控制周期", _make_sinusoidal
    ),
    ValueStrategy("random_walk", "布朗运动，当前值基础上随机游走", _make_random_walk),
    ValueStrategy("ramp", "线性斜坡，到 hi 后回 lo", _make_ramp),
    ValueStrategy("counter", "每次 tick 累加，到 hi 后回 lo", _make_counter),
    ValueStrategy("constant", "恒定值，从 gen_params.value 取", _make_constant),
    ValueStrategy("random_current", "电机电流仿真：60% 中心 + 正弦抖动", _make_current),
    ValueStrategy("random_temp", "温度仿真：40% 中心 + 小抖动", _make_temp),
    ValueStrategy("random_pressure", "压力仿真：50% 中心 + 微小抖动", _make_pressure),
    ValueStrategy("random_speed", "转速仿真：70% 中心 + 微小抖动", _make_speed),
]


class StrategyRegistry:
    """Holds all known value-generation strategies and picks the right one."""

    def __init__(self, strategies: Optional[List[ValueStrategy]] = None):
        self._map: Dict[str, ValueStrategy] = {}
        for s in strategies or _BUILTIN_STRATEGIES:
            self.register(s)

    def register(self, strategy: ValueStrategy) -> None:
        self._map[strategy.name] = strategy

    def get(self, name: str) -> Optional[ValueStrategy]:
        return self._map.get(name)

    def list_all(self) -> List[Dict[str, str]]:
        return [
            {"name": s.name, "description": s.description} for s in self._map.values()
        ]

    # ------------------------------------------------------------------
    # Auto-selection
    # ------------------------------------------------------------------

    _AUTO_MAP: Dict[str, str] = {
        "bool": "binary_toggle",
        "string": "constant",
        "int": "random_uniform",
        "float": "random_uniform",
    }

    def resolve(self, node: NodeMeta) -> ValueStrategy:
        """Return the effective strategy for *node*.

        Resolution order:
        1. Explicit ``gen_strategy`` if set.
        2. ``data_type`` → sensible default.
        3. Fallback: ``random_uniform``.
        """
        name = node.gen_strategy or self._AUTO_MAP.get(node.data_type, "random_uniform")
        return self._map.get(name, self._map["random_uniform"])

    # ------------------------------------------------------------------
    # Generate a value for one node
    # ------------------------------------------------------------------

    def generate(
        self,
        node: NodeMeta,
        current_value: float,
        elapsed_sec: float = 0.0,
        tick: int = 0,
    ) -> Any:
        strategy = self.resolve(node)
        raw = strategy.generate(
            node.range_lo,
            node.range_hi,
            node.gen_params,
            current_value,
            elapsed_sec,
            tick,
        )

        if node.data_type == "bool":
            return 1 if raw >= 0.5 else 0
        if node.data_type == "int":
            return int(round(raw))
        if node.data_type == "string":
            return str(raw)
        # float
        return round(raw, 3)


# ============================================================================
# Model parsing – Excel → DeviceModel
# ============================================================================

# Column-name aliases so users can write either "node_id" or "NodeId" etc.
_COLUMN_ALIASES: Dict[str, str] = {
    "node_id": "node_id",
    "nodeid": "node_id",
    "node id": "node_id",
    "data_type": "data_type",
    "datatype": "data_type",
    "data type": "data_type",
}

_OPTIONAL_COLUMNS = {
    "range_lo": ["range_lo", "rangelo", "低限", "下限", "min", "min_value"],
    "range_hi": ["range_hi", "rangehi", "高限", "上限", "max", "max_value"],
    "unit": ["unit", "单位", "engineering_units"],
    "gen_strategy": ["gen_strategy", "strategy", "策略", "生成策略"],
    "gen_params": ["gen_params", "params", "策略参数"],
    "description": ["description", "描述", "备注", "说明"],
    "instance_count": ["instance_count", "实例数", "数量", "count"],
    "display_name": ["display_name", "displayname", "显示名", "别名"],
    "group_depth": ["group_depth", "分组深度"],
}

_REQUIRED_COLUMNS = ["node_id", "data_type"]


def _normalize_header(raw: str) -> str:
    """Lower-case and strip a header cell."""
    return str(raw).strip().lower() if raw else ""


def _build_column_map(headers: List[str]) -> Dict[str, int]:
    """Map normalized header names → column index (0-based)."""
    col_map: Dict[str, int] = {}
    norm_headers = [_normalize_header(h) for h in headers]

    # Required columns — exact or alias match
    for req in _REQUIRED_COLUMNS:
        aliases = [req] + [a for a, target in _COLUMN_ALIASES.items() if target == req]
        for idx, nh in enumerate(norm_headers):
            if nh in aliases:
                col_map[req] = idx
                break
        if req not in col_map:
            raise ValueError(
                f"Missing required column '{req}'.  Found headers: {headers}"
            )

    # Optional columns — try each alias
    for col_field, aliases in _OPTIONAL_COLUMNS.items():
        for alias in aliases:
            for idx, nh in enumerate(norm_headers):
                if nh == alias:
                    col_map[col_field] = idx
                    break
            if col_field in col_map:
                break

    return col_map


def _safe_str(val: Any) -> str:
    return str(val).strip() if val is not None else ""


def _safe_float(val: Any, default: float = 0.0) -> float:
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _safe_int(val: Any, default: int = 1) -> int:
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _safe_json(val: Any) -> Dict[str, Any]:
    if val is None:
        return {}
    s = str(val).strip()
    if not s:
        return {}
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        # Try parsing key=value or key:value lines
        result: Dict[str, Any] = {}
        for part in s.replace(";", ",").split(","):
            part = part.strip()
            if not part:
                continue
            for sep in (":", "="):
                if sep in part:
                    k, v = part.split(sep, 1)
                    try:
                        result[k.strip()] = float(v.strip())
                    except ValueError:
                        result[k.strip()] = v.strip()
                    break
            else:
                result[part] = True
        return result


def read_model_excel(
    xlsx_path: str | Path = "opc_sim_list.xlsx",
    sheet_name: Optional[str] = None,
) -> Tuple[List[NodeMeta], List[str]]:
    """Parse a generic OPC model spreadsheet.

    Returns
    -------
    (nodes, warnings)
        List of :class:`NodeMeta` rows and any non-fatal warnings.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = sheet_name or wb.sheetnames[0]
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ValueError("Spreadsheet is empty — no header row found.")

    headers = [str(c) if c is not None else "" for c in header_row]
    col = _build_column_map(headers)

    nodes: List[NodeMeta] = []
    warnings: List[str] = []

    for row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(c is None for c in row):
            continue

        node_id = (
            _safe_str(row[col["node_id"]])
            if "node_id" in col and len(row) > col["node_id"]
            else ""
        )
        if not node_id:
            warnings.append(f"Row {row_num}: empty node_id, skipping")
            continue

        data_type = (
            _safe_str(row[col["data_type"]]).lower()
            if "data_type" in col and len(row) > col["data_type"]
            else "float"
        )
        if data_type not in ("float", "int", "bool", "string"):
            warnings.append(
                f"Row {row_num}: unknown data_type '{data_type}', defaulting to float"
            )
            data_type = "float"

        def _opt(field: str, default: Any = "") -> Any:
            if field not in col:
                return default
            idx = col[field]
            return row[idx] if idx < len(row) and row[idx] is not None else default

        meta = NodeMeta(
            node_id=node_id,
            data_type=data_type,
            range_lo=_safe_float(_opt("range_lo"), 0.0),
            range_hi=_safe_float(_opt("range_hi"), 100.0),
            unit=_safe_str(_opt("unit")),
            gen_strategy=_safe_str(_opt("gen_strategy")),
            gen_params=_safe_json(_opt("gen_params")),
            description=_safe_str(_opt("description")),
            instance_count=_safe_int(_opt("instance_count"), 1),
            display_name=_safe_str(_opt("display_name")),
            group_depth=_safe_int(_opt("group_depth"), 1),
        )

        if meta.range_lo > meta.range_hi:
            warnings.append(f"Row {row_num}: range_lo > range_hi, swapped")
            meta.range_lo, meta.range_hi = meta.range_hi, meta.range_lo

        nodes.append(meta)

    wb.close()
    return nodes, warnings


def build_device_model(nodes: List[NodeMeta]) -> DeviceModel:
    """Build a complete :class:`DeviceModel` from a list of :class:`NodeMeta`.

    Expands ``instance_count`` and groups nodes by ``group_key``.
    """
    model = DeviceModel()
    group_order: List[str] = []

    for meta in nodes:
        gk = meta.group_key
        if gk not in model.groups:
            model.groups[gk] = GroupInfo(key=gk, label=gk)
            group_order.append(gk)

        if meta.instance_count <= 1:
            model.nodes[meta.node_id] = meta
            model.groups[gk].node_ids.append(meta.node_id)
            model.groups[gk].node_count += 1
        else:
            for i in range(meta.instance_count):
                inst_id = meta.instance_node_id(i)
                inst_meta = NodeMeta(
                    node_id=inst_id,
                    data_type=meta.data_type,
                    range_lo=meta.range_lo,
                    range_hi=meta.range_hi,
                    unit=meta.unit,
                    gen_strategy=meta.gen_strategy,
                    gen_params=dict(meta.gen_params),
                    description=meta.description,
                    instance_count=1,
                    display_name=f"{meta.effective_display_name}_{i:03d}",
                    group_depth=meta.group_depth,
                )
                model.nodes[inst_id] = inst_meta
                model.groups[gk].node_ids.append(inst_id)
                model.groups[gk].node_count += 1

    return model


# ============================================================================
# State persistence – JSON snapshot of runtime changes
# ============================================================================


def node_meta_to_dict(node: NodeMeta) -> Dict[str, Any]:
    """Serialize NodeMeta to a plain dict (safe for JSON)."""
    d = asdict(node)
    d["gen_params"] = json.dumps(node.gen_params, ensure_ascii=False, default=str)
    return d


def dict_to_node_meta(d: Dict[str, Any]) -> NodeMeta:
    """Deserialize from a plain dict back to NodeMeta."""
    d = dict(d)  # shallow copy
    if "gen_params" in d and isinstance(d["gen_params"], str):
        try:
            d["gen_params"] = json.loads(d["gen_params"])
        except json.JSONDecodeError:
            d["gen_params"] = {}
    return NodeMeta(**{k: d[k] for k in d if k in NodeMeta.__dataclass_fields__})


def save_model_state(model: DeviceModel, json_path: str | Path) -> None:
    """Persist the current (potentially Web-edited) model state to JSON.

    Atomic write: serialize to a temp file in the same directory, then rename
    over the target.  A crash mid-write will not corrupt the existing state.
    """
    data = {
        "nodes": {nid: node_meta_to_dict(meta) for nid, meta in model.nodes.items()},
        "group_order": list(model.groups.keys()),
        "separator": model.separator,
    }
    path = Path(json_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
    # tempfile + os.replace is atomic on both POSIX and Windows (Python 3.3+)
    fd, tmp_path = tempfile.mkstemp(
        prefix=".tmp_" + path.name + "_",
        suffix=".json",
        dir=str(path.parent),
    )
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(payload)
        os.replace(tmp_path, str(path))
    except Exception:
        # Best-effort cleanup of the temp file on failure
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def load_model_state(json_path: str | Path) -> Optional[DeviceModel]:
    """Load a previously saved model state.

    Returns ``None`` if the file doesn't exist.
    """
    path = Path(json_path)
    if not path.exists():
        return None
    data = json.loads(path.read_text(encoding="utf-8"))
    model = DeviceModel(separator=data.get("separator", "."))
    for nid, d in data.get("nodes", {}).items():
        node = dict_to_node_meta(d)
        node.node_id = nid  # ensure consistency
        model.nodes[nid] = node
    # Rebuild groups from node data
    for nid in data.get("group_order", []):
        if nid in model.nodes:
            gk = model.nodes[nid].group_key
            if gk not in model.groups:
                model.groups[gk] = GroupInfo(key=gk, label=gk)
            model.groups[gk].node_ids.append(nid)
            model.groups[gk].node_count += 1
    # Catch any nodes not in group_order
    for nid, meta in model.nodes.items():
        gk = meta.group_key
        if gk not in model.groups:
            model.groups[gk] = GroupInfo(key=gk, label=gk)
        if nid not in model.groups[gk].node_ids:
            model.groups[gk].node_ids.append(nid)
            model.groups[gk].node_count += 1
    return model


# ============================================================================
# Singleton – Strategy registry
# ============================================================================

_strategy_registry: Optional[StrategyRegistry] = None


def get_strategy_registry() -> StrategyRegistry:
    global _strategy_registry
    if _strategy_registry is None:
        _strategy_registry = StrategyRegistry()
    return _strategy_registry
