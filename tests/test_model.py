"""Tests for Excel parsing and JSON state persistence in ``common.py``.

Covers:
- read_model_excel: required columns, optional columns, aliases, type coercion
- build_device_model: instance expansion, group derivation
- node_meta_to_dict / dict_to_node_meta: JSON round-trip
- save_model_state / load_model_state: file persistence (atomic write)
- Edge cases: empty sheet, missing columns, range_lo > range_hi swap
"""

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from common import (
    DeviceModel,
    GroupInfo,
    NodeMeta,
    build_device_model,
    dict_to_node_meta,
    load_model_state,
    node_meta_to_dict,
    read_model_excel,
    save_model_state,
)


# ── Fixtures ────────────────────────────────────────────────────────


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """Create a small sample xlsx with the expected schema."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PointTable"
    ws.append(
        [
            "node_id",
            "data_type",
            "range_lo",
            "range_hi",
            "unit",
            "gen_strategy",
            "gen_params",
        ]
    )
    ws.append(
        [
            "Shearer.left_motor.current",
            "float",
            0,
            100,
            "A",
            "random_current",
            '{"center_ratio":0.6}',
        ]
    )
    ws.append(["Shearer.left_motor.switch", "bool", 0, 1, "", "binary_toggle", ""])
    ws.append(["Belt.main.speed", "int", 0, 1500, "rpm", "random_uniform", ""])
    path = tmp_path / "sample.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def empty_xlsx(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(["node_id", "data_type"])  # header only, no data rows
    path = tmp_path / "empty.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def missing_col_xlsx(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(["node_id"])  # missing data_type column
    ws.append(["some.node"])
    path = tmp_path / "bad.xlsx"
    wb.save(str(path))
    return path


# ── read_model_excel ────────────────────────────────────────────────


def test_read_model_excel_parses_rows(sample_xlsx: Path):
    nodes, warnings = read_model_excel(str(sample_xlsx))
    assert len(nodes) == 3
    assert warnings == []

    n0 = nodes[0]
    assert n0.node_id == "Shearer.left_motor.current"
    assert n0.data_type == "float"
    assert n0.range_lo == 0.0
    assert n0.range_hi == 100.0
    assert n0.unit == "A"
    assert n0.gen_strategy == "random_current"
    assert n0.gen_params == {"center_ratio": 0.6}


def test_read_model_excel_bool_and_int_types(sample_xlsx: Path):
    nodes, _ = read_model_excel(str(sample_xlsx))
    bool_node = next(n for n in nodes if n.data_type == "bool")
    int_node = next(n for n in nodes if n.data_type == "int")
    assert bool_node.node_id == "Shearer.left_motor.switch"
    assert int_node.node_id == "Belt.main.speed"


def test_read_model_excel_empty_returns_no_nodes(empty_xlsx: Path):
    nodes, warnings = read_model_excel(str(empty_xlsx))
    assert nodes == []
    assert warnings == []


def test_read_model_excel_missing_required_column_raises(missing_col_xlsx: Path):
    with pytest.raises(ValueError, match="Missing required column"):
        read_model_excel(str(missing_col_xlsx))


def test_read_model_excel_unknown_data_type_warns_and_defaults(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["node_id", "data_type"])
    ws.append(["x.y", "weird_type"])
    p = tmp_path / "weird.xlsx"
    wb.save(str(p))
    nodes, warnings = read_model_excel(str(p))
    assert len(nodes) == 1
    assert nodes[0].data_type == "float"  # defaulted
    assert any("unknown data_type" in w for w in warnings)


def test_read_model_excel_range_swap(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["node_id", "data_type", "range_lo", "range_hi"])
    ws.append(["x.y", "float", 100, 0])  # lo > hi
    p = tmp_path / "swap.xlsx"
    wb.save(str(p))
    nodes, warnings = read_model_excel(str(p))
    assert nodes[0].range_lo == 0.0
    assert nodes[0].range_hi == 100.0
    assert any("swapped" in w for w in warnings)


def test_read_model_excel_empty_node_id_skipped(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["node_id", "data_type"])
    ws.append(["", "float"])  # empty node_id
    ws.append(["good.node", "float"])
    p = tmp_path / "empty_id.xlsx"
    wb.save(str(p))
    nodes, warnings = read_model_excel(str(p))
    assert len(nodes) == 1
    assert nodes[0].node_id == "good.node"
    assert any("empty node_id" in w for w in warnings)


# ── build_device_model ──────────────────────────────────────────────


def test_build_device_model_groups_by_first_segment(sample_xlsx: Path):
    nodes, _ = read_model_excel(str(sample_xlsx))
    model = build_device_model(nodes)
    assert "Shearer" in model.groups
    assert "Belt" in model.groups
    assert len(model.groups["Shearer"].node_ids) == 2
    assert len(model.groups["Belt"].node_ids) == 1


def test_build_device_model_instance_expansion():
    meta = NodeMeta(node_id="A.B.C", data_type="float", instance_count=3)
    model = build_device_model([meta])
    assert len(model.nodes) == 3
    assert "A.B.C_000" in model.nodes
    assert "A.B.C_001" in model.nodes
    assert "A.B.C_002" in model.nodes
    # Each instance has instance_count=1 (no further recursion)
    for n in model.nodes.values():
        assert n.instance_count == 1


# ── node_meta_to_dict / dict_to_node_meta round-trip ─────────────────


def test_node_meta_round_trip_preserves_fields():
    original = NodeMeta(
        node_id="Shearer.motor.current",
        data_type="float",
        range_lo=0.0,
        range_hi=100.0,
        unit="A",
        gen_strategy="random_current",
        gen_params={"center_ratio": 0.6, "jitter_ratio": 0.15},
        description="Main motor current",
        instance_count=1,
        display_name="MotorCurrent",
        group_depth=1,
    )
    d = node_meta_to_dict(original)
    # gen_params is serialized to a string in the dict form
    assert isinstance(d["gen_params"], str)
    parsed = json.loads(d["gen_params"])
    assert parsed == {"center_ratio": 0.6, "jitter_ratio": 0.15}

    restored = dict_to_node_meta(d)
    assert restored.node_id == original.node_id
    assert restored.data_type == original.data_type
    assert restored.range_lo == original.range_lo
    assert restored.range_hi == original.range_hi
    assert restored.unit == original.unit
    assert restored.gen_strategy == original.gen_strategy
    assert restored.gen_params == original.gen_params  # dict again
    assert restored.description == original.description
    assert restored.display_name == original.display_name


def test_dict_to_node_meta_handles_invalid_gen_params_string():
    d = {"node_id": "x", "gen_params": "not valid json"}
    meta = dict_to_node_meta(d)
    assert meta.gen_params == {}  # falls back to empty dict


def test_dict_to_node_meta_drops_unknown_keys():
    d = {"node_id": "x", "unknown_field": "should be dropped", "data_type": "float"}
    meta = dict_to_node_meta(d)
    assert meta.node_id == "x"
    assert meta.data_type == "float"
    assert not hasattr(meta, "unknown_field")


# ── save_model_state / load_model_state ─────────────────────────────


def test_save_and_load_model_state_round_trip(tmp_path: Path):
    model = DeviceModel()
    model.nodes["A.B.temp"] = NodeMeta(
        node_id="A.B.temp", data_type="float", range_lo=0, range_hi=100, unit="C"
    )
    model.nodes["A.B.switch"] = NodeMeta(node_id="A.B.switch", data_type="bool")
    model.groups["A"] = GroupInfo(
        key="A", label="A", node_ids=["A.B.temp", "A.B.switch"], node_count=2
    )

    state_path = tmp_path / "state.json"
    save_model_state(model, state_path)
    assert state_path.exists()

    loaded = load_model_state(state_path)
    assert loaded is not None
    assert set(loaded.nodes.keys()) == {"A.B.temp", "A.B.switch"}
    assert loaded.nodes["A.B.temp"].unit == "C"
    assert loaded.nodes["A.B.switch"].data_type == "bool"
    assert "A" in loaded.groups
    assert len(loaded.groups["A"].node_ids) == 2


def test_save_model_state_no_temp_leftovers(tmp_path: Path):
    """Atomic write must clean up the temp file even though it succeeds."""
    model = DeviceModel()
    model.nodes["x"] = NodeMeta(node_id="x")
    model.groups["x"] = GroupInfo(key="x", label="x", node_ids=["x"], node_count=1)
    state_path = tmp_path / "state.json"
    save_model_state(model, state_path)
    leftovers = list(tmp_path.glob(".tmp_*"))
    assert leftovers == []


def test_load_model_state_missing_file_returns_none(tmp_path: Path):
    assert load_model_state(tmp_path / "does_not_exist.json") is None


def test_save_model_state_overwrites_existing(tmp_path: Path):
    """Atomic write via os.replace should overwrite, not append."""
    state_path = tmp_path / "state.json"
    state_path.write_text('{"old": true}', encoding="utf-8")

    model = DeviceModel()
    model.nodes["x"] = NodeMeta(node_id="x")
    model.groups["x"] = GroupInfo(key="x", label="x", node_ids=["x"], node_count=1)
    save_model_state(model, state_path)

    # Should be valid JSON with our structure, not the old content
    data = json.loads(state_path.read_text(encoding="utf-8"))
    assert "nodes" in data
    assert "old" not in data
