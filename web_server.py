"""
OPC 模拟器 Web 控制台
提供 Web 界面来控制 OPC 节点的数值
运行方式: python web_server.py
"""

import json
import math
import os
import random
import threading
import time
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
from pathlib import Path

import openpyxl
from opcua import Server, ua

# React 构建产物目录
DIST_DIR = Path(__file__).parent / 'frontend' / 'dist'


# ─── 点表读取 ────────────────────────────────────────────────────────────────

def read_point_table(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Sheet1']
    equipment = [
        (1, 2, 'control', '设备启停'),
        (3, 4, 'shearer', '采煤机'),
        (5, 6, 'support', '电液控'),
        (7, 8, 'sanji', '三机'),
        (9, 10, 'belt', '皮带'),
        (11, 12, 'liquid', '泵站'),
        (13, 14, 'power', '供电'),
        (15, 16, 'transformer', '移变'),
        (17, 18, 'switch', '开关'),
        (19, 20, 'alarm', '故障'),
    ]
    all_points = []
    for name_col, addr_col, key, label in equipment:
        points = []
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=name_col).value
            if name is None:
                continue
            points.append(str(name))
        all_points.append((key, label, points))
    return all_points


# ─── 模拟值生成 ──────────────────────────────────────────────────────────────

def generate_value(point_name, tick):
    name = point_name.lower()
    if '通讯' in name: return 1
    if '运行' in name or '状态' in name: return 1 if tick % 20 < 15 else 0
    if '请求开' in name: return 1 if tick % 30 == 0 else 0
    if '请求关' in name: return 1 if tick % 30 == 15 else 0
    if '闭锁' in name or '急停' in name: return 0
    if '远控' in name or '自动' in name or '手动' in name: return 1
    if '电流' in name:
        base = 50.0 if '采煤机' in name else 30.0
        return round(base + random.uniform(-5, 5) + 2 * math.sin(tick * 0.1), 2)
    if '电压' in name: return round(3300 + random.uniform(-50, 50), 1)
    if '温度' in name: return round(45 + random.uniform(-5, 15) + 3 * math.sin(tick * 0.05), 1)
    if '压力' in name: return round(20 + random.uniform(-3, 5) + 2 * math.sin(tick * 0.08), 2)
    if '转速' in name: return round(1450 + random.uniform(-50, 50), 0)
    if '采高' in name or '高度' in name: return round(3.5 + random.uniform(-0.3, 0.3), 2)
    if '位置' in name or '架号' in name: return int(50 + 20 * math.sin(tick * 0.02))
    if '速度' in name: return round(5 + random.uniform(-1, 3), 1)
    if '浓度' in name: return round(3.5 + random.uniform(-0.5, 0.5), 1)
    if '液位' in name or '油位' in name or '水位' in name: return round(60 + random.uniform(-10, 10), 1)
    if '瓦斯' in name: return round(0.05 + random.uniform(0, 0.15), 3)
    if '倾角' in name or '俯仰角' in name: return round(random.uniform(-5, 5), 1)
    if '转向' in name or '向左' in name or '向右' in name or '方向' in name: return random.choice([0, 1])
    if '工作方式' in name: return random.choice([0, 1, 2])
    if '开机率' in name: return round(70 + random.uniform(-10, 20), 1)
    if '时间' in name: return int(tick * 60 + random.randint(0, 3600))
    if '保护' in name: return 0
    if '通断' in name: return 1
    if '故障' in name: return 0
    if '张力' in name: return round(50 + random.uniform(-5, 5), 1)
    if '频率' in name: return round(50 + random.uniform(-1, 1), 2)
    if '功率' in name: return round(100 + random.uniform(-20, 30), 1)
    if '转矩' in name: return round(80 + random.uniform(-10, 15), 1)
    if '湿度' in name: return round(40 + random.uniform(-5, 15), 1)
    if '跟机' in name: return random.choice([0, 1])
    return round(random.uniform(0, 100), 2)


# ─── OPC 模拟器 ──────────────────────────────────────────────────────────────

class MiningOPCSimulator:
    def __init__(self, endpoint='opc.tcp://0.0.0.0:14840/freeopcua/server/'):
        self.server = Server()
        self.server.set_endpoint(endpoint)
        self.server.set_server_name('Mining Equipment Simulator')
        self.nodes = {}
        self.point_names = []
        self.modes = {}       # {name: 'random' | 'manual'}
        self.manual_vals = {} # {name: float}
        self.current_vals = {}
        self.tick = 0
        self.running = True

    def setup(self, excel_path):
        all_points = read_point_table(excel_path)
        uri = 'http://mining.simulator'
        idx = self.server.register_namespace(uri)
        objects = self.server.get_objects_node()

        for key, label, points in all_points:
            folder = objects.add_folder(idx, label)
            for i, name in enumerate(points):
                node = folder.add_variable(
                    ua.NodeId(f'{key}/{i}', idx), name, 0.0, ua.VariantType.Double
                )
                node.set_writable()
                # 使用 folder_key/index 作为唯一键，避免重复点名覆盖
                unique_key = f'{key}/{i}'
                self.nodes[unique_key] = (name, node)
                self.point_names.append(name)
                self.modes[unique_key] = 'random'
                self.manual_vals[unique_key] = 0.0
                self.current_vals[unique_key] = 0.0

    def update_values(self):
        for unique_key, (name, node) in self.nodes.items():
            try:
                if self.modes[unique_key] == 'manual':
                    val = self.manual_vals[unique_key]
                else:
                    val = generate_value(name, self.tick)
                node.set_value(val)
                self.current_vals[unique_key] = val
            except Exception:
                pass
        self.tick += 1

    def run(self):
        self.server.start()
        print(f'OPC 服务器已启动: {self.server.endpoint}')
        try:
            while self.running:
                self.update_values()
                time.sleep(2)
        except KeyboardInterrupt:
            pass
        finally:
            self.server.stop()


# ─── Web API ─────────────────────────────────────────────────────────────────

simulator = None

# MIME 类型映射
MIME_TYPES = {
    '.html': 'text/html; charset=utf-8',
    '.js': 'application/javascript; charset=utf-8',
    '.css': 'text/css; charset=utf-8',
    '.json': 'application/json; charset=utf-8',
    '.png': 'image/png',
    '.jpg': 'image/jpeg',
    '.svg': 'image/svg+xml',
    '.ico': 'image/x-icon',
    '.woff': 'font/woff',
    '.woff2': 'font/woff2',
}


def read_dist_file(relative_path):
    """读取 dist 目录下的文件"""
    file_path = DIST_DIR / relative_path.lstrip('/')
    if file_path.exists() and file_path.is_file():
        return file_path.read_bytes()
    return None


class APIHandler(SimpleHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path

        # API 路由
        if path.startswith('/api/'):
            self.handle_api_get(path)
            return

        # 静态文件服务
        if path == '/' or path == '/index.html':
            content = read_dist_file('index.html')
            if content:
                self.send_response(200)
                self.send_header('Content-Type', 'text/html; charset=utf-8')
                self.end_headers()
                self.wfile.write(content)
                return
            self.send_error(404, 'index.html not found')
            return

        # 尝试从 dist 目录提供静态文件
        content = read_dist_file(path)
        if content:
            ext = Path(path).suffix
            mime = MIME_TYPES.get(ext, 'application/octet-stream')
            self.send_response(200)
            self.send_header('Content-Type', mime)
            self.end_headers()
            self.wfile.write(content)
            return

        # SPA 回退：所有非 API/静态文件请求返回 index.html
        content = read_dist_file('index.html')
        if content:
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(content)
            return

        self.send_error(404)

    def handle_api_get(self, path):
        if path == '/api/nodes':
            result = []
            for key, label, points in read_point_table(
                os.path.join(os.path.dirname(os.path.abspath(__file__)), 'opc_list_test.xlsx')
            ):
                group = {'key': key, 'label': label, 'nodes': []}
                for i, name in enumerate(points):
                    unique_key = f'{key}/{i}'
                    group['nodes'].append({
                        'name': name,
                        'mode': simulator.modes.get(unique_key, 'random'),
                        'value': simulator.current_vals.get(unique_key, 0),
                        'manual': simulator.manual_vals.get(unique_key, 0),
                    })
                result.append(group)
            self.send_json(result)
            return

        if path == '/api/values':
            self.send_json({
                unique_key: simulator.current_vals.get(unique_key, 0)
                for unique_key in simulator.current_vals.keys()
            })
            return

        self.send_error(404)

    def do_POST(self):
        parsed = urlparse(self.path)
        body = json.loads(self.rfile.read(int(self.headers['Content-Length'])))

        if parsed.path == '/api/set_mode':
            unique_key = body.get('unique_key') or body.get('name')
            mode = body.get('mode')
            if unique_key in simulator.modes:
                simulator.modes[unique_key] = mode
                self.send_json({'ok': True})
            else:
                self.send_error(404, 'Node not found')
            return

        if parsed.path == '/api/set_value':
            unique_key = body.get('unique_key') or body.get('name')
            value = body.get('value')
            if unique_key in simulator.manual_vals:
                simulator.manual_vals[unique_key] = value
                simulator.modes[unique_key] = 'manual'
                self.send_json({'ok': True})
            else:
                self.send_error(404, 'Node not found')
            return

        if parsed.path == '/api/set_all_mode':
            mode = body.get('mode')
            for unique_key in simulator.modes:
                simulator.modes[unique_key] = mode
            self.send_json({'ok': True})
            return

        self.send_error(404)

    def send_json(self, data):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))

    def log_message(self, format, *args):
        pass  # 静默日志


# ─── 启动 ────────────────────────────────────────────────────────────────────

def main():
    global simulator

    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, 'opc_list_test.xlsx')

    simulator = MiningOPCSimulator()
    simulator.setup(excel_path)
    print(f'已加载 {len(simulator.point_names)} 个节点')

    opc_thread = threading.Thread(target=simulator.run, daemon=True)
    opc_thread.start()

    web_port = 18480
    httpd = HTTPServer(('0.0.0.0', web_port), APIHandler)
    print(f'Web 控制台: http://localhost:{web_port}')
    print(f'OPC 服务器: opc.tcp://localhost:14840')
    print(f'按 Ctrl+C 停止')

    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        simulator.running = False
        httpd.shutdown()
        print('已停止')


if __name__ == '__main__':
    main()
